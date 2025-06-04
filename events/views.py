from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.views import View
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.http import HttpResponse
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Q
from urllib.parse import quote
import csv
from io import BytesIO
import qrcode

from .models import Event, EventField, Registration
from .forms import EventForm, EventFieldForm, DynamicRegistrationForm

class EventListView(LoginRequiredMixin, ListView):
    model = Event
    template_name = 'events/event_list.html'
    context_object_name = 'events'
    paginate_by = 10

    def get_queryset(self):
        return Event.objects.filter(organizer=self.request.user).order_by('-created_at')

class EventDetailView(LoginRequiredMixin, DetailView):
    model = Event
    template_name = 'events/event_detail.html'
    context_object_name = 'event'

    def get_queryset(self):
        return Event.objects.filter(organizer=self.request.user)

class EventCreateView(LoginRequiredMixin, CreateView):
    model = Event
    form_class = EventForm
    template_name = 'events/event_form.html'

    def form_valid(self, form):
        form.instance.organizer = self.request.user
        messages.success(self.request, 'Event created successfully!')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('events:form_builder', kwargs={'slug': self.object.slug})

class EventUpdateView(LoginRequiredMixin, UpdateView):
    model = Event
    form_class = EventForm
    template_name = 'events/event_form.html'

    def get_queryset(self):
        return Event.objects.filter(organizer=self.request.user)

    def form_valid(self, form):
        messages.success(self.request, 'Event updated successfully!')
        return super().form_valid(form)

class EventDeleteView(LoginRequiredMixin, DeleteView):
    model = Event
    template_name = 'events/event_confirm_delete.html'
    success_url = reverse_lazy('events:event_list')

    def get_queryset(self):
        return Event.objects.filter(organizer=self.request.user)

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Event deleted successfully!')
        return super().delete(request, *args, **kwargs)

class EventFormBuilderView(LoginRequiredMixin, DetailView):
    model = Event
    template_name = 'events/form_builder.html'
    context_object_name = 'event'

    def get_queryset(self):
        return Event.objects.filter(organizer=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fields'] = self.object.fields.all().order_by('order', 'id')
        context['field_form'] = EventFieldForm()
        return context

class AddEventFieldView(LoginRequiredMixin, View):
    def post(self, request, slug):
        event = get_object_or_404(Event, slug=slug, organizer=request.user)
        form = EventFieldForm(request.POST)
        
        if form.is_valid():
            field = form.save(commit=False)
            field.event = event
            field.save()
            messages.success(request, 'Field added successfully!')
        else:
            messages.error(request, 'Error adding field. Please check the form.')
        
        return redirect('events:form_builder', slug=slug)

class EditEventFieldView(LoginRequiredMixin, View):
    def post(self, request, slug, field_id):
        event = get_object_or_404(Event, slug=slug, organizer=request.user)
        field = get_object_or_404(EventField, id=field_id, event=event)
        form = EventFieldForm(request.POST, instance=field)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Field updated successfully!')
        else:
            messages.error(request, 'Error updating field.')
        
        return redirect('events:form_builder', slug=slug)

class DeleteEventFieldView(LoginRequiredMixin, View):
    def post(self, request, slug, field_id):
        event = get_object_or_404(Event, slug=slug, organizer=request.user)
        field = get_object_or_404(EventField, id=field_id, event=event)
        field.delete()
        messages.success(request, 'Field deleted successfully!')
        return redirect('events:form_builder', slug=slug)

class EventRegistrationView(View):
    def get(self, request, slug):
        event = get_object_or_404(Event, slug=slug, is_published=True)
        form = DynamicRegistrationForm(event)
        return render(request, 'events/registration_form.html', {
            'event': event,
            'form': form
        })

    def post(self, request, slug):
        event = get_object_or_404(Event, slug=slug, is_published=True)
        form = DynamicRegistrationForm(event, request.POST, request.FILES)
        
        if form.is_valid():
            registration = form.save()
              # Send email with QR code (optional)
            if registration.participant_email:
                self.send_confirmation_email(registration)
            
            return redirect('events:registration_success', unique_id=registration.unique_id)
        
        return render(request, 'events/registration_form.html', {
            'event': event,
            'form': form
        })

    def send_confirmation_email(self, registration):
        try:
            subject = f'Registration Confirmation - {registration.event.name}'
            message = f'''
            Dear {registration.participant_name},
            
            Thank you for registering for {registration.event.name}!
            
            Event Details:
            - Date: {registration.event.start_date}
            - Time: {registration.event.start_time}
            - Location: {registration.event.location}
            
            Your QR Code: {self.request.build_absolute_uri(reverse('events:qr_code', kwargs={'unique_id': registration.unique_id}))}
            
            Please bring this QR code to the event for check-in.
            
            Best regards,
            {registration.event.organizer.get_full_name() or registration.event.organizer.username}
            '''
            
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [registration.participant_email],
                fail_silently=True,
            )
            
            # Log email sending in development mode
            if settings.EMAIL_BACKEND == 'django.core.mail.backends.console.EmailBackend':
                print(f"[DEVELOPMENT MODE] Confirmation email logged to console for: {registration.participant_email}")
            else:
                print(f"Confirmation email sent to: {registration.participant_email}")
                
        except Exception as e:
            print(f"Email sending failed: {e}")

class RegistrationSuccessView(DetailView):
    model = Registration
    template_name = 'events/registration_success.html'
    context_object_name = 'registration'

    def get_object(self):
        return get_object_or_404(Registration, unique_id=self.kwargs['unique_id'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registration = self.get_object()
        
        # Generate QR code URL
        qr_code_url = self.request.build_absolute_uri(
            reverse('events:qr_code', kwargs={'unique_id': registration.unique_id})
        )
        
        # Create WhatsApp sharing text
        whatsapp_text = f"Here's my registration for {registration.event.name}: {qr_code_url}"
        
        # WhatsApp sharing URL
        context['whatsapp_share_url'] = f"https://wa.me/?text={quote(whatsapp_text)}"
        
        return context

class QRCodeView(DetailView):
    model = Registration
    template_name = 'events/qr_code.html'
    context_object_name = 'registration'

    def get_object(self):
        return get_object_or_404(Registration, unique_id=self.kwargs['unique_id'])

class ParticipantListView(LoginRequiredMixin, DetailView):
    model = Event
    template_name = 'events/participant_list.html'
    context_object_name = 'event'

    def get_queryset(self):
        return Event.objects.filter(organizer=self.request.user)
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registrations = Registration.objects.filter(event=self.object).order_by('-registered_at')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            registrations = registrations.filter(
                Q(participant_name__icontains=search) |
                Q(participant_email__icontains=search) |
                Q(participant_phone__icontains=search)
            )
        
        # Status filter
        status = self.request.GET.get('status')
        if status:
            registrations = registrations.filter(status=status)
        
        # Count statistics
        context['total_registrations'] = registrations.count()
        context['attended_count'] = registrations.filter(status='attended').count()
        
        context['registrations'] = registrations
        context['search'] = search
        context['status'] = status
        return context

class ExportParticipantsView(LoginRequiredMixin, View):
    def get(self, request, pk):
        event = get_object_or_404(Event, pk=pk, organizer=request.user)
        format_type = request.GET.get('format', 'csv')
        
        if format_type == 'excel':
            return self.export_excel(event)
        else:
            return self.export_csv(event)
    
    def export_csv(self, event):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{event.slug}_participants.csv"'
        
        writer = csv.writer(response)
          # Header row
        headers = ['Name', 'Email', 'Phone', 'Status', 'Registered At', 'Attended At']
        # Add custom field headers
        custom_fields = event.fields.all().order_by('order')
        headers.extend([field.field_name for field in custom_fields])
        writer.writerow(headers)
        
        # Data rows
        for registration in event.registrations.all():
            row = [
                registration.participant_name,
                registration.participant_email,
                registration.participant_phone or '',
                registration.get_status_display(),
                registration.registered_at.strftime('%Y-%m-%d %H:%M'),
                registration.attended_at.strftime('%Y-%m-%d %H:%M') if registration.attended_at else '',
            ]
            
            # Add custom field data
            field_data = registration.get_field_data()
            for field in custom_fields:
                row.append(field_data.get(field.field_name, ''))
            
            writer.writerow(row)
        
        return response
    
    def export_excel(self, event):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"{event.name} Participants"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
          # Header row
        headers = ['Name', 'Email', 'Phone', 'Status', 'Registered At', 'Attended At']
        custom_fields = event.fields.all().order_by('order')
        headers.extend([field.field_name for field in custom_fields])
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row_num, registration in enumerate(event.registrations.all(), 2):
            data = [
                registration.participant_name,
                registration.participant_email,
                registration.participant_phone or '',
                registration.get_status_display(),
                registration.registered_at.strftime('%Y-%m-%d %H:%M'),
                registration.attended_at.strftime('%Y-%m-%d %H:%M') if registration.attended_at else '',
            ]
              # Add custom field data
            field_data = registration.get_field_data()
            for field in custom_fields:
                data.append(field_data.get(field.field_name, ''))
            
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{event.slug}_participants.xlsx"'
        
        return response

class PublishEventView(LoginRequiredMixin, View):
    def post(self, request, pk):
        event = get_object_or_404(Event, pk=pk, organizer=request.user)
        event.is_published = True
        event.save()
        messages.success(request, f'Event "{event.name}" has been published successfully!')
        return redirect('events:event_detail', pk=event.pk)

class EventQRCodeView(LoginRequiredMixin, View):
    def get(self, request, pk):
        event = get_object_or_404(Event, pk=pk, organizer=request.user)
        
        response = HttpResponse(content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="{event.slug}_qr_code.png"'
        
        # Generate QR code image
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(request.build_absolute_uri(event.registration_url))
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        img_io = BytesIO()
        img.save(img_io, 'PNG')
        img_io.seek(0)
        
        response.write(img_io.getvalue())
        return response

class RegistrationDetailView(LoginRequiredMixin, DetailView):
    model = Registration
    template_name = 'events/registration_detail.html'
    context_object_name = 'registration'

    def get_queryset(self):
        return Registration.objects.filter(event__organizer=self.request.user)

class RegistrationQRCodeView(DetailView):
    model = Registration
    template_name = 'events/registration_qr.html'
    context_object_name = 'registration'

    def get_object(self):
        return get_object_or_404(Registration, pk=self.kwargs['pk'])
